import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\thescoop\coding\rateConvert\Time & Fees.xlsx')
ws = wb['Table 1']

print(f'Total rows in Excel: {ws.max_row}')
print('\nFirst 80 rows:\n')

for i in range(1, min(81, ws.max_row+1)):
    date_val = ws.cell(i, 1).value
    status_val = ws.cell(i, 2).value
    type_val = ws.cell(i, 3).value
    desc_val = ws.cell(i, 4).value
    
    if date_val or status_val:
        print(f'Row {i}: Date={date_val}, Status={status_val}, Type={type_val}, Desc={str(desc_val)[:30] if desc_val else ""}')
