import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import glob
from datetime import datetime
import sys


class RateConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Woodruff Billing - Legal Aid Rate Converter (c) 2025")
        self.root.geometry("1100x600")
        self.root.resizable(True, True)
        
        self.rates_file = None
        self.rates_data = None
        self.input_file = None
        self.loaded_data = None
        self.converted_wb = None
        self.converted_ws = None
        self.conversion_result = None
        self.current_theme = 'light'
        self.style = ttk.Style()
        self.preview_mode = 'review'  # 'review', 'input', or 'output'
        self.input_rows = []
        self.output_rows = []
        self.rows_for_review = []
        
        self.setup_ui()
        self.apply_light_theme()
        self.discover_rates_file()
    
    def apply_light_theme(self):
        self.current_theme = 'light'
        self.style.theme_use('clam')
        
        # Main background - soft grey
        self.root.configure(bg='#E8EBF0')
        
        # Configure frame styles with depth
        self.style.configure('TFrame', background='#E8EBF0')
        self.style.configure('Card.TFrame', background='#FFFFFF', relief='flat', borderwidth=0)
        
        # Configure LabelFrame (section headers) - Clean white cards
        self.style.configure('TLabelframe', background='#FFFFFF', borderwidth=0, relief='flat')
        self.style.configure('TLabelframe.Label', background='#FFFFFF', foreground='#1A1A1A', 
                       font=('Segoe UI', 9, 'bold'), padding=(10, 6))
        
        # Configure Labels
        self.style.configure('TLabel', background='#E8EBF0', foreground='#1A1A1A', 
                       font=('Segoe UI', 8))
        self.style.configure('Title.TLabel', background='#E8EBF0', foreground='#1A1A1A', 
                       font=('Segoe UI', 13, 'bold'))
        
        # Configure Buttons - compact but substantial
        self.style.configure('TButton', background='#17A2B8', foreground='white', 
                       font=('Segoe UI', 8, 'bold'), padding=(8, 6), relief='flat', 
                       borderwidth=0)
        self.style.map('TButton', background=[('active', '#138496'), ('disabled', '#D0D0D0')])
        
        # Convert button (blue)
        self.style.configure('Convert.TButton', background='#0066CC', foreground='white', 
                       font=('Segoe UI', 9, 'bold'), padding=(10, 8), relief='flat')
        self.style.map('Convert.TButton', background=[('active', '#0052A3'), ('disabled', '#D0D0D0')])
        
        # Save button (green)
        self.style.configure('Save.TButton', background='#28A745', foreground='white', 
                       font=('Segoe UI', 9, 'bold'), padding=(10, 8), relief='flat')
        self.style.map('Save.TButton', background=[('active', '#218838'), ('disabled', '#D0D0D0')])
        
        # Theme toggle button - compact icon
        self.style.configure('Theme.TButton', background='#6C757D', foreground='white', 
                       font=('Segoe UI', 12), padding=(6, 4), relief='flat')
        self.style.map('Theme.TButton', background=[('active', '#5A6268')])
        
        # Combobox
        self.style.configure('TCombobox', fieldbackground='white', background='#17A2B8', 
                       foreground='#1A1A1A', borderwidth=0, relief='flat')
        
        # Treeview - compact
        self.style.configure('Treeview', background='white', foreground='#1A1A1A', 
                       fieldbackground='white', font=('Segoe UI', 8), rowheight=24)
        self.style.configure("Treeview.Heading", 
                       background="#4A6FA5",
                       foreground="white",
                       font=('Segoe UI', 8, 'bold'),
                       relief="flat",
                       borderwidth=0)
        self.style.map("Treeview.Heading",
                 background=[('active', '#5B7FC7')],
                 foreground=[('active', 'white')])
        
        # Update status text
        if hasattr(self, 'status_text'):
            self.status_text.config(bg='#2C3E50', fg='#00FF00')
        
        # Update drop zone
        if hasattr(self, 'drop_zone'):
            self.drop_zone.config(bg='#F5F5F5', relief='solid', bd=1)
            if hasattr(self, 'drop_icon_label'):
                self.drop_icon_label.config(background='#F5F5F5')
                self.drop_text_label.config(background='#F5F5F5')
                self.drop_hint_label.config(background='#F5F5F5')
        
        # Update stats labels
        if hasattr(self, 'total_rows_label'):
            pass  # TLabel style already applied
    
    def apply_dark_theme(self):
        self.current_theme = 'dark'
        self.style.theme_use('clam')
        
        # Main background - Softer dark blue
        self.root.configure(bg='#1E2939')
        
        # Configure frame styles with depth
        self.style.configure('TFrame', background='#1E2939')
        self.style.configure('Card.TFrame', background='#2A3F54', relief='flat', borderwidth=0)
        
        # Configure LabelFrame (section headers) - Dark cards
        self.style.configure('TLabelframe', background='#2A3F54', borderwidth=0, relief='flat')
        self.style.configure('TLabelframe.Label', background='#2A3F54', foreground='#E8EBF0', 
                       font=('Segoe UI', 9, 'bold'), padding=(10, 6))
        
        # Configure Labels
        self.style.configure('TLabel', background='#1E2939', foreground='#E8EBF0', 
                       font=('Segoe UI', 8))
        self.style.configure('Title.TLabel', background='#1E2939', foreground='#E8EBF0', 
                       font=('Segoe UI', 13, 'bold'))
        
        # Configure Buttons - compact but substantial
        self.style.configure('TButton', background='#17A2B8', foreground='white', 
                       font=('Segoe UI', 8, 'bold'), padding=(8, 6), relief='flat', 
                       borderwidth=0)
        self.style.map('TButton', background=[('active', '#138496'), ('disabled', '#3A3A3A')])
        
        # Convert button (bright blue)
        self.style.configure('Convert.TButton', background='#007BFF', foreground='white', 
                       font=('Segoe UI', 9, 'bold'), padding=(10, 8), relief='flat')
        self.style.map('Convert.TButton', background=[('active', '#0056B3'), ('disabled', '#3A3A3A')])
        
        # Save button (bright green)
        self.style.configure('Save.TButton', background='#28A745', foreground='white', 
                       font=('Segoe UI', 9, 'bold'), padding=(10, 8), relief='flat')
        self.style.map('Save.TButton', background=[('active', '#1E7E34'), ('disabled', '#3A3A3A')])
        
        # Theme toggle button - compact icon
        self.style.configure('Theme.TButton', background='#6C757D', foreground='white', 
                       font=('Segoe UI', 12), padding=(6, 4), relief='flat')
        self.style.map('Theme.TButton', background=[('active', '#5A6268')])
        
        # Combobox - Light background with dark text for readability
        self.style.configure('TCombobox', fieldbackground='#ECF0F1', background='#17A2B8', 
                       foreground='#000000', borderwidth=0, relief='flat')
        self.style.map('TCombobox', fieldbackground=[('readonly', '#ECF0F1')], 
                      foreground=[('readonly', '#000000')])
        
        # Treeview - Dark mode compact
        self.style.configure('Treeview', background='#2A3F54', foreground='#E8EBF0', 
                       fieldbackground='#2A3F54', font=('Segoe UI', 8), rowheight=24)
        self.style.configure("Treeview.Heading", 
                       background="#17A2B8",
                       foreground="white",
                       font=('Segoe UI', 8, 'bold'),
                       relief="flat",
                       borderwidth=0)
        self.style.map("Treeview.Heading",
                 background=[('active', '#1ABC9C')],
                 foreground=[('active', 'white')])
        
        # Update status text
        if hasattr(self, 'status_text'):
            self.status_text.config(bg='#1C2833', fg='#00FF00')
        
        # Update drop zone
        if hasattr(self, 'drop_zone'):
            self.drop_zone.config(bg='#3A4F64', relief='solid', bd=1)
            if hasattr(self, 'drop_icon_label'):
                self.drop_icon_label.config(background='#3A4F64')
                self.drop_text_label.config(background='#3A4F64')
                self.drop_hint_label.config(background='#3A4F64')
        
        # Update stats labels
        if hasattr(self, 'total_rows_label'):
            pass  # TLabel style already applied
    
    def toggle_theme(self):
        if self.current_theme == 'light':
            self.apply_dark_theme()
            self.theme_btn.config(text="🌙")
        else:
            self.apply_light_theme()
            self.theme_btn.config(text="☀")
    
    def setup_ui(self):
        paned_window = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        left_frame = ttk.Frame(paned_window, width=400)
        right_frame = ttk.Frame(paned_window, width=800)
        
        paned_window.add(left_frame, weight=1)
        paned_window.add(right_frame, weight=2)
        
        self.setup_left_panel(left_frame)
        self.setup_right_panel(right_frame)
    
    def setup_left_panel(self, parent):
        parent.configure(style='TFrame')
        
        title_label = ttk.Label(parent, text="Legal Aid Rate Converter", 
                                style='Title.TLabel')
        title_label.pack(pady=(6, 10))
        
        ttk.Separator(parent, orient='horizontal').pack(fill='x', pady=3)
        
        # 1. Rates Reference File
        rates_frame = ttk.LabelFrame(parent, text="Rates Reference File", padding="6")
        rates_frame.pack(fill='x', padx=10, pady=6)
        
        self.rates_status_label = ttk.Label(rates_frame, text="⚪ No rates file loaded", 
                                           foreground="gray")
        self.rates_status_label.pack(anchor='w')
        
        self.rates_path_label = ttk.Label(rates_frame, text="", 
                                         foreground="gray", font=('Arial', 8))
        self.rates_path_label.pack(anchor='w', pady=(2, 5))
        
        change_rates_btn = ttk.Button(rates_frame, text="Change Rates File...", 
                                     command=self.browse_rates_file)
        change_rates_btn.pack(pady=5)
        
        # 2. Excel File to Convert
        input_frame = ttk.LabelFrame(parent, text="Excel File to Convert", padding="6")
        input_frame.pack(fill='x', padx=10, pady=6)
        
        # Drop zone frame
        self.drop_zone = tk.Frame(input_frame, height=60, relief='flat', 
                                 bd=2, cursor='hand2')
        self.drop_zone.pack(fill='both', expand=True)
        self.drop_zone.pack_propagate(False)
        
        # Icon and text
        self.drop_icon_label = ttk.Label(self.drop_zone, text="📊", font=('Segoe UI', 16))
        self.drop_icon_label.pack(pady=(5, 0))
        
        self.drop_text_label = ttk.Label(self.drop_zone, text="Drop 'Time and Fees.xlsx' here",
                                        font=('Segoe UI', 9, 'bold'))
        self.drop_text_label.pack()
        
        self.drop_hint_label = ttk.Label(self.drop_zone, text="(or click to browse)",
                                        font=('Segoe UI', 7), foreground="gray")
        self.drop_hint_label.pack()
        
        # Bind click event
        self.drop_zone.bind('<Button-1>', lambda e: self.browse_input_file())
        self.drop_icon_label.bind('<Button-1>', lambda e: self.browse_input_file())
        self.drop_text_label.bind('<Button-1>', lambda e: self.browse_input_file())
        self.drop_hint_label.bind('<Button-1>', lambda e: self.browse_input_file())
        
        # Setup drag and drop if available
        if HAS_DND:
            self.drop_zone.drop_target_register(DND_FILES)
            self.drop_zone.dnd_bind('<<Drop>>', self.on_file_drop)
            self.drop_zone.dnd_bind('<<DragEnter>>', self.on_drag_enter)
            self.drop_zone.dnd_bind('<<DragLeave>>', self.on_drag_leave)
        
        # 3. Target Case Type
        case_frame = ttk.LabelFrame(parent, text="Target Case Type", padding="6")
        case_frame.pack(fill='x', padx=10, pady=6)
        
        self.case_type_var = tk.StringVar()
        self.case_type_combo = ttk.Combobox(case_frame, textvariable=self.case_type_var, 
                                           state='readonly', width=35)
        self.case_type_combo.pack(fill='x')
        
        # 4. Action Buttons
        buttons_frame = ttk.Frame(parent)
        buttons_frame.pack(fill='x', padx=10, pady=8)
        
        self.convert_btn = ttk.Button(buttons_frame, text="Convert Rates", 
                                     command=self.convert_rates, state='disabled',
                                     style='Convert.TButton')
        self.convert_btn.pack(side='left', padx=5, expand=True, fill='x')
        
        self.save_btn = ttk.Button(buttons_frame, text="Save Output", 
                                  command=self.save_output, state='disabled',
                                  style='Save.TButton')
        self.save_btn.pack(side='left', padx=5, expand=True, fill='x')
        
        # 5. Status Log
        status_frame = ttk.LabelFrame(parent, text="Status Log", padding="5")
        status_frame.pack(fill='both', expand=True, padx=10, pady=6)
        
        self.status_text = scrolledtext.ScrolledText(status_frame, height=15, 
                                                    wrap=tk.WORD, state='disabled',
                                                    font=('Consolas', 8),
                                                    bg='#2C3E50', fg='#00FF00',
                                                    insertbackground='white')
        self.status_text.pack(fill='both', expand=True)
        
        self.status_text.tag_config('success', foreground='#00FF00')
        self.status_text.tag_config('warning', foreground='#FFA500')
        self.status_text.tag_config('error', foreground='#FF4444')
        self.status_text.tag_config('info', foreground='#00BFFF')
        self.status_text.tag_config('normal', foreground='#CCCCCC')
    
    def setup_right_panel(self, parent):
        header_frame = ttk.Frame(parent)
        header_frame.pack(fill='x', pady=(10, 5), padx=10)
        
        self.preview_title_label = ttk.Label(header_frame, text="Preview: For Review", 
                               font=('Segoe UI', 11, 'bold'))
        self.preview_title_label.pack(side='left')
        
        self.theme_btn = ttk.Button(header_frame, text="☀", command=self.toggle_theme, 
                              width=3, style='Theme.TButton')
        self.theme_btn.pack(side='right', padx=5)
        
        help_btn = ttk.Button(header_frame, text="Help", command=self.show_help, width=8)
        help_btn.pack(side='right', padx=5)
        
        # Single toggle button for Input/Output
        self.view_toggle_btn = ttk.Button(header_frame, text="View: Input ▼", 
                                         command=self.toggle_input_output, width=15)
        self.view_toggle_btn.pack(side='right', padx=5)
        self.view_toggle_btn.config(state='disabled')  # Disabled until conversion runs
        
        # Use PanedWindow to split table (2/3) and stats (1/3) vertically
        content_pane = tk.PanedWindow(parent, orient=tk.VERTICAL, sashwidth=3, 
                                     sashrelief=tk.RAISED, bg='#cccccc')
        content_pane.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Table section (top 2/3)
        tree_container = ttk.Frame(content_pane)
        content_pane.add(tree_container, stretch='always')
        
        tree_frame = ttk.Frame(tree_container)
        tree_frame.pack(fill='both', expand=True)
        
        columns = ('Row#', 'Date', 'Status', 'Type', 'Description', 'Staff', 'Hrs/Qty', 'Amount', 'Issue')
        self.preview_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=20)
        
        column_widths_chars = {
            'Row#': 50,
            'Date': 11 * 7, 
            'Status': 20 * 7, 
            'Type': 7 * 7, 
            'Description': 45 * 7, 
            'Staff': 18 * 7, 
            'Hrs/Qty': 10 * 7, 
            'Amount': 10 * 7, 
            'Issue': 120
        }
        
        for col in columns:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=column_widths_chars.get(col, 100), anchor='w')
        
        self.preview_tree.tag_configure('missing_data', background='#FFFF99')
        self.preview_tree.tag_configure('unmatched', background='#FFD699')
        self.preview_tree.tag_configure('success_message', background='#D4EDDA', font=('Arial', 10, 'bold'))
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.preview_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.preview_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Stats section (bottom 1/3)
        stats_container = ttk.Frame(content_pane)
        content_pane.add(stats_container, stretch='always')
        
        stats_frame = ttk.LabelFrame(stats_container, text="Conversion Statistics", padding="10")
        stats_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Row 1: Counts
        counts_frame = ttk.Frame(stats_frame)
        counts_frame.pack(fill='x', pady=2)
        
        self.total_rows_label = ttk.Label(counts_frame, text="Total: 0", font=('Segoe UI', 8))
        self.total_rows_label.pack(side='left', padx=(0, 15))
        
        self.converted_rows_label = ttk.Label(counts_frame, text="Converted: 0", font=('Segoe UI', 8))
        self.converted_rows_label.pack(side='left', padx=(0, 15))
        
        self.unmatched_rows_label = ttk.Label(counts_frame, text="Unmatched: 0", font=('Segoe UI', 8))
        self.unmatched_rows_label.pack(side='left')
        
        # Row 2: Totals
        totals_row_frame = ttk.Frame(stats_frame)
        totals_row_frame.pack(fill='x', pady=2)
        
        self.original_total_label = ttk.Label(totals_row_frame, text="Original Total: £0.00", 
                                             font=('Segoe UI', 8))
        self.original_total_label.pack(side='left', padx=(0, 15))
        
        self.converted_total_label = ttk.Label(totals_row_frame, text="Converted Total: £0.00", 
                                              font=('Segoe UI', 8, 'bold'), foreground='green')
        self.converted_total_label.pack(side='left', padx=(0, 15))
        
        self.difference_label = ttk.Label(totals_row_frame, text="Difference: £0.00", 
                                         font=('Segoe UI', 8), foreground='gray')
        self.difference_label.pack(side='left')
        
        # Store reference to PanedWindow for later positioning
        self.content_pane = content_pane
        
        # Set initial sash position after window is shown (2/3 split)
        self.root.after(100, lambda: self.content_pane.sash_place(0, 0, int(self.content_pane.winfo_height() * 0.66)))
    
    def log_status(self, message, tag='normal'):
        self.status_text.config(state='normal')
        self.status_text.insert(tk.END, message + '\n', tag)
        self.status_text.see(tk.END)
        self.status_text.config(state='disabled')
        self.root.update()
    
    def update_preview(self, rows, issue_type):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        if not rows:
            self.preview_tree.insert('', 'end', values=('', '', '', '', '✓ Analysis Complete', '', '', '', ''), 
                                    tags=('success_message',))
            self.preview_tree.insert('', 'end', values=('', '', '', '', 'No Issues Found - Ready to Convert', '', '', '', ''), 
                                    tags=('success_message',))
            return
        
        tag = 'missing_data' if issue_type == 'missing' else 'unmatched'
        
        for row in rows:
            date_str = row['date'].strftime('%d/%m/%Y') if isinstance(row['date'], datetime) else str(row['date'])
            row_num = row.get('row_num', '')
            values = (
                row_num,
                date_str,
                row.get('status', ''),
                row.get('type', ''),
                row.get('description', ''),
                row.get('staff', ''),
                row.get('hrs_qty', ''),
                row.get('amount', ''),
                row.get('issue', '')
            )
            self.preview_tree.insert('', 'end', values=values, tags=(tag,))
    
    def show_review_view(self):
        self.preview_mode = 'review'
        self.preview_title_label.config(text="Preview: For Review", foreground='orange')
        self.view_toggle_btn.config(state='disabled')
        self.display_review_rows()
    
    def toggle_input_output(self):
        if self.preview_mode == 'input':
            # Switch to output
            self.preview_mode = 'output'
            self.preview_title_label.config(text="Preview: OUTPUT (Converted Rates)", foreground='green')
            self.view_toggle_btn.config(text="View: Output ▼")
            self.display_output_rows()
        else:
            # Switch to input (from output or review)
            self.preview_mode = 'input'
            self.preview_title_label.config(text="Preview: INPUT (Original Rates)", foreground='blue')
            self.view_toggle_btn.config(text="View: Input ▼")
            self.display_input_rows()
    
    def display_input_rows(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        if not self.input_rows:
            return
        
        for row in self.input_rows:
            date_str = row['date'].strftime('%d/%m/%Y') if isinstance(row['date'], datetime) else str(row['date']) if row.get('date') else ''
            values = (
                row.get('row_num', ''),
                date_str,
                row.get('status', ''),
                row.get('type', ''),
                row.get('description', ''),
                row.get('staff', ''),
                row.get('hrs_qty', ''),
                row.get('amount', ''),
                ''
            )
            self.preview_tree.insert('', 'end', values=values)
    
    def display_output_rows(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        if not self.output_rows:
            return
        
        for row in self.output_rows:
            date_str = row['date'].strftime('%d/%m/%Y') if isinstance(row['date'], datetime) else str(row['date']) if row.get('date') else ''
            tag = 'unmatched' if not row.get('matched', True) else ''
            values = (
                row.get('row_num', ''),
                date_str,
                row.get('status', ''),
                row.get('type', ''),
                row.get('description', ''),
                row.get('staff', ''),
                row.get('hrs_qty', ''),
                row.get('new_amount', row.get('amount', '')),
                'Unmatched' if tag else ''
            )
            self.preview_tree.insert('', 'end', values=values, tags=(tag,) if tag else ())
    
    def display_review_rows(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        if not self.rows_for_review:
            # Show success message if no rows to review
            self.preview_tree.insert('', 'end', values=('', '', '', '', '✓ All rows converted successfully - no issues found', '', '', '', ''), 
                                    tags=('success_message',))
            return
        
        for row in self.rows_for_review:
            date_str = row['date'].strftime('%d/%m/%Y') if isinstance(row['date'], datetime) else str(row['date']) if row.get('date') else ''
            tag = 'unmatched' if not row.get('matched', True) else ''
            values = (
                row.get('row_num', ''),
                date_str,
                row.get('status', ''),
                row.get('type', ''),
                row.get('description', ''),
                row.get('staff', ''),
                row.get('hrs_qty', ''),
                row.get('new_amount', row.get('amount', '')),
                row.get('issue', '')
            )
            self.preview_tree.insert('', 'end', values=values, tags=('unmatched',))
    
    def update_totals(self, original_total, converted_total):
        difference = converted_total - original_total
        
        # Update totals
        self.original_total_label.config(text=f"Original Total: £{original_total:,.2f}")
        self.converted_total_label.config(text=f"Converted Total: £{converted_total:,.2f}")
        
        diff_color = "green" if difference >= 0 else "red"
        diff_sign = "+" if difference >= 0 else ""
        self.difference_label.config(
            text=f"Difference: {diff_sign}£{difference:,.2f}", 
            foreground=diff_color
        )
    
    def update_stats(self, total, converted, unmatched):
        self.total_rows_label.config(text=f"Total: {total}")
        self.converted_rows_label.config(text=f"Converted: {converted}")
        self.unmatched_rows_label.config(text=f"Unmatched: {unmatched}", 
                                        foreground='orange' if unmatched > 0 else 'green')
    
    def show_help(self):
        help_window = tk.Toplevel(self.root)
        help_window.title("Legal Aid Rate Converter - Help")
        help_window.geometry("700x600")
        help_window.resizable(False, False)
        
        help_text = scrolledtext.ScrolledText(help_window, wrap=tk.WORD, font=('Arial', 10), padx=20, pady=20)
        help_text.pack(fill='both', expand=True)
        
        help_content = """LEGAL AID RATE CONVERTER - USER GUIDE

═══════════════════════════════════════════════════════════

HOW TO USE THIS APPLICATION

1. RATES REFERENCE FILE
   The program automatically searches for a clearbill file containing Legal Aid rates:
   • First checks: OneDrive\_latest_scripts folder
   • Then checks: Desktop
   • Uses the most recently modified file if multiple found
   • Click "Change Rates File..." to manually select a different file

2. SELECT TARGET CASE TYPE
   Choose which Legal Aid case type rates you want to convert TO:
   • Care & Supervision
   • Private Law Family
   • Other Public Law
   • etc.

3. SELECT INPUT FILE
   Click "Select File..." to choose the Time & Fees Excel file you want to convert.
   This should be an Excel file exported from LEAP or Osprey with "Table 1" sheet.

4. ANALYZE INPUT (Optional)
   Click "Analyze Input" to check for issues BEFORE converting:
   • Missing Hrs/Qty values (highlighted in YELLOW)
   • Missing Amount values (highlighted in YELLOW)
   If issues found, fix them in Excel and reload the file.

5. CONVERT RATES
   Click "Convert Rates" to perform the conversion:
   • Calculates current rate for each row (Amount ÷ Hrs/Qty)
   • Matches rate to activity type (Attendance, Letters Out, etc.)
   • Recalculates amount using target case type rates
   • Conversion is done IN MEMORY - not saved yet

6. REVIEW PREVIEW PANEL
   After conversion, the right panel shows "Rows for Review":
   • ORANGE HIGHLIGHTED rows = Unmatched rates (couldn't auto-convert)
   • Common reasons for unmatched rows:
     - FAS (Family Advocacy Scheme) entries - rates vary by bundle size
     - Custom rates not in standard Legal Aid rates table
     - Missing or zero Hrs/Qty values
   
   FAS entries will typically show as unmatched - this is NORMAL.
   You can manually correct FAS amounts in Excel after saving.

7. CHECK TOTALS
   Review the "Totals Comparison" section:
   • Original Total: Sum of all amounts in input file
   • Converted Total: Sum after rate conversion
   • Difference: Shows increase/decrease (GREEN if positive, RED if negative)

8. SAVE OUTPUT
   When satisfied with the conversion:
   • Click "Save Output" to write changes to the Excel file
   • If file is open in Excel, you'll be prompted to close it first
   • Original data is preserved in "orig_timeFees" sheet
   • Converted data is saved to "Table 1" sheet
   • Unmatched rows are highlighted in YELLOW in the saved file

═══════════════════════════════════════════════════════════

UNDERSTANDING THE PREVIEW PANEL

What rows appear here?
• BEFORE conversion (Analyze): Rows with missing Hrs/Qty or Amount
• AFTER conversion (Convert): Rows that couldn't be auto-converted

Why are FAS entries shown?
FAS (Family Advocacy Scheme) rates are not in the standard rates table because 
they vary based on court bundle size. These will appear as "unmatched" and 
should be manually reviewed/corrected as needed.

Color coding:
• YELLOW background = Missing data (before conversion)
• ORANGE background = Unmatched rate (after conversion, inc. FAS)
• "✓ No problems found" = All rows converted successfully (rare with FAS)

═══════════════════════════════════════════════════════════

COMMON WORKFLOWS

Converting standard Legal Aid work:
1. Select input file → Convert Rates → Review (FAS will show) → Save Output
2. Manually update FAS amounts in Excel if needed

Checking for data quality issues:
1. Select input file → Analyze Input → Fix issues in Excel → Reload → Convert

Converting with custom rates:
1. Convert Rates → Review unmatched rows
2. Manually update custom-rate rows in Excel after saving

═══════════════════════════════════════════════════════════

TIPS

• FAS entries showing as unmatched is NORMAL - handle them manually
• Always check the Totals Comparison to verify overall conversion
• Original data is always preserved in "orig_timeFees" sheet
• Close Excel file before clicking "Save Output" to avoid errors
• Status Log (left panel) shows detailed progress and any errors

═══════════════════════════════════════════════════════════
"""
        
        help_text.insert('1.0', help_content)
        help_text.config(state='disabled')
        
        close_btn = ttk.Button(help_window, text="Close", command=help_window.destroy)
        close_btn.pack(pady=10)
    
    def discover_rates_file(self):
        self.log_status("Searching for rates file...", 'info')
        
        onedrive_dir = r"C:\Users\thescoop\OneDrive\_latest_scripts"
        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        
        search_locations = [
            (onedrive_dir, "OneDrive (_latest_scripts)"),
            (desktop_dir, "Desktop")
        ]
        
        for location, location_name in search_locations:
            if os.path.exists(location):
                pattern = os.path.join(location, "*clearbill*.xlsm")
                found_files = glob.glob(pattern)
                
                if found_files:
                    most_recent = max(found_files, key=os.path.getmtime)
                    self.rates_file = most_recent
                    self.log_status(f"✓ Found rates file in {location_name}", 'success')
                    self.load_rates_file(self.rates_file)
                    return
        
        self.log_status("⚠ No clearbill file found in OneDrive or Desktop", 'warning')
        self.log_status("Please select rates file manually using 'Change Rates File' button", 'info')
        self.rates_status_label.config(text="⚠ No rates file found", foreground="orange")
    
    def browse_rates_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Clearbill (BOC Excel) file for rates",
            filetypes=[("Excel Macro Files", "*.xlsm"), ("All Files", "*.*")],
            initialdir=r"C:\Users\thescoop\OneDrive\_latest_scripts"
        )
        
        if file_path:
            self.rates_file = file_path
            self.load_rates_file(file_path)
    
    def load_rates_file(self, file_path):
        try:
            self.log_status(f"Loading rates from: {os.path.basename(file_path)}", 'info')
            
            wb = openpyxl.load_workbook(file_path, keep_vba=True, data_only=True)
            
            if 'Rates' not in wb.sheetnames:
                raise ValueError("Selected file does not contain 'Rates' sheet")
            
            ws = wb['Rates']
            
            case_types = []
            for col in range(6, ws.max_column + 1):
                header = ws.cell(7, col).value
                if header and header != '< Select >':
                    case_types.append(header.strip().replace('\n', ' '))
            
            activity_mapping = {
                'Attendance (or Preparation)': 'attend',
                'Preparation (same as above)': 'attend',
                'Advocacy (not used here)': 'advocacy',
                'Attendance behind counsel': 'wCounsel',
                'Travel (or Waiting)': 'travelWait',
                'Waiting (same as above)': 'travelWait',
                'Letters Out': 'lOut',
                'Letters In': 'lIn',
                'Phone Calls': 'call'
            }
            
            rates = {}
            
            for row in range(8, 17):
                activity_label = ws.cell(row, 4).value
                if activity_label and activity_label in activity_mapping:
                    activity_key = activity_mapping[activity_label]
                    
                    if activity_key not in rates:
                        rates[activity_key] = []
                        
                        for col in range(6, 6 + len(case_types)):
                            rate_value = ws.cell(row, col).value
                            if rate_value and isinstance(rate_value, (int, float)):
                                rates[activity_key].append(float(rate_value))
                            else:
                                rates[activity_key].append(0.0)
            
            wb.close()
            
            self.rates_data = {
                'case_types': case_types,
                'rates': rates
            }
            
            self.case_type_combo['values'] = case_types
            if case_types:
                self.case_type_combo.current(0)
            
            self.rates_status_label.config(text=f"✓ {os.path.basename(file_path)}", 
                                          foreground="green")
            self.rates_path_label.config(text=file_path)
            
            self.log_status(f"✓ Loaded {len(case_types)} case types and {len(rates)} activity types", 'success')
            
            self.check_ready_to_analyze()
            
        except Exception as e:
            messagebox.showerror("Error Loading Rates File", str(e))
            self.log_status(f"✗ Error loading rates file: {str(e)}", 'error')
    
    def browse_input_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Time & Fees Excel file to convert",
            filetypes=[("Excel Files", "*.xlsx;*.xlsm"), ("All Files", "*.*")],
            initialdir=os.getcwd()
        )
        
        if file_path:
            self.load_input_file(file_path)
    
    def load_input_file(self, file_path):
        if not os.path.exists(file_path):
            messagebox.showerror("File Not Found", f"File does not exist:\n{file_path}")
            return
        
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.xlsx', '.xlsm']:
            messagebox.showerror("Invalid File", "Please select an Excel file (.xlsx or .xlsm)")
            return
        
        self.input_file = file_path
        filename = os.path.basename(file_path)
        
        # Update drop zone to show selected file
        self.drop_icon_label.config(text="✓")
        self.drop_text_label.config(text=filename)
        self.drop_hint_label.config(text="Click to change file")
        self.reset_drop_zone_style()
        
        self.log_status(f"Selected input file: {filename}", 'info')
        self.check_ready_to_analyze()
        self.analyze_input()
    
    def on_file_drop(self, event):
        file_path = event.data
        # Remove curly braces if present (Windows DND format)
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        # Remove quotes if present
        file_path = file_path.strip('"')
        
        self.load_input_file(file_path)
        self.on_drag_leave(None)
    
    def on_drag_enter(self, event):
        self.drop_zone.config(relief='solid', bd=2)
    
    def on_drag_leave(self, event):
        self.reset_drop_zone_style()
    
    def reset_drop_zone_style(self):
        if self.current_theme == 'light':
            self.drop_zone.config(bg='#F5F5F5', relief='solid', bd=1)
        else:
            self.drop_zone.config(bg='#3A4F64', relief='solid', bd=1)
    
    def check_ready_to_analyze(self):
        if self.rates_data and self.input_file:
            self.convert_btn.config(state='normal')
        else:
            self.convert_btn.config(state='disabled')
    
    def analyze_input(self):
        self.log_status("\nAnalyzing input file for issues...", 'info')
        
        try:
            wb, ws, data_rows = self.load_time_fees_table(self.input_file)
            wb.close()
            
            # Store input rows for preview
            self.input_rows = data_rows
            
            problematic_rows = []
            
            for row in data_rows:
                # Skip completely empty rows
                is_empty_row = not any([
                    row.get('date'),
                    row.get('description'),
                    row.get('amount'),
                    row.get('hrs_qty'),
                    row.get('staff')
                ])
                
                if is_empty_row:
                    continue
                
                hrs_qty = row.get('hrs_qty')
                amount = row.get('amount')
                
                # Check for missing hrs/qty
                hrs_qty_missing = (hrs_qty is None or hrs_qty == '' or 
                                  (isinstance(hrs_qty, (int, float)) and hrs_qty == 0))
                
                # Check for missing amount
                amount_missing = (amount is None or amount == '' or 
                                 (isinstance(amount, (int, float)) and amount == 0))
                
                if hrs_qty_missing:
                    row['issue'] = 'Missing Hrs/Qty'
                    problematic_rows.append(row)
                elif amount_missing:
                    row['issue'] = 'Missing Amount'
                    problematic_rows.append(row)
            
            # Store for preview
            self.rows_for_review = problematic_rows
            
            if problematic_rows:
                self.log_status(f"⚠ Found {len(problematic_rows)} row(s) with missing data", 'warning')
                self.update_preview(problematic_rows, 'missing')
            else:
                self.log_status("✓ No issues found in input file", 'success')
                self.update_preview([], 'missing')
            
        except Exception as e:
            self.log_status(f"✗ Error analyzing input: {str(e)}", 'error')
            messagebox.showerror("Analysis Error", str(e))
    
    def convert_rates(self):
        self.log_status("\n" + "="*60, 'info')
        self.log_status("Starting conversion...", 'info')
        self.log_status("="*60, 'info')
        
        try:
            target_case_type = self.case_type_var.get()
            target_index = self.rates_data['case_types'].index(target_case_type)
            
            self.log_status(f"Target case type: {target_case_type}", 'info')
            self.log_status(f"Input file: {os.path.basename(self.input_file)}", 'info')
            
            wb, ws, data_rows = self.load_time_fees_table(self.input_file)
            
            original_total = 0.0
            for row in data_rows:
                amount_val = self.parse_amount(row['amount'])
                if amount_val:
                    original_total += amount_val
            
            conversion_result = self.perform_conversion(data_rows, target_index)
            
            converted_total = 0.0
            for row in conversion_result['converted_rows']:
                amount_val = self.parse_amount(row['new_amount'])
                if amount_val:
                    converted_total += amount_val
            
            self.write_output(wb, ws, conversion_result, target_case_type)
            
            self.converted_wb = wb
            self.converted_ws = ws
            self.conversion_result = conversion_result
            
            # Store data for toggle views
            self.output_rows = conversion_result['converted_rows']
            
            # Build rows for review - only unmatched rows with actual data
            unmatched_rows = []
            for row in conversion_result['converted_rows']:
                if not row.get('matched', True):
                    row['issue'] = 'Unmatched rate'
                    unmatched_rows.append(row)
            
            self.rows_for_review = unmatched_rows
            
            # Update stats and totals
            self.update_totals(original_total, converted_total)
            self.update_stats(len(data_rows), len(data_rows) - conversion_result['unmatched_count'], 
                            conversion_result['unmatched_count'])
            
            self.log_status("\n" + "="*60, 'success')
            self.log_status("CONVERSION COMPLETE (not saved yet)", 'success')
            self.log_status("="*60, 'success')
            self.log_status(f"Total rows: {len(data_rows)}", 'info')
            self.log_status(f"Converted: {len(data_rows) - conversion_result['unmatched_count']}", 'success')
            self.log_status(f"Unmatched: {conversion_result['unmatched_count']}", 'warning' if conversion_result['unmatched_count'] > 0 else 'info')
            self.log_status(f"Original total: £{original_total:,.2f}", 'info')
            self.log_status(f"Converted total: £{converted_total:,.2f}", 'success')
            self.log_status(f"Difference: £{converted_total - original_total:,.2f}", 'info')
            
            # Show review view by default
            self.show_review_view()
            
            # Enable the Input/Output toggle button
            self.view_toggle_btn.config(state='normal')
            
            if conversion_result['unmatched_count'] > 0:
                self.log_status("\n⚠ Review unmatched rows before saving", 'warning')
            else:
                self.log_status("✓ All rows converted successfully", 'success')
            
            self.save_btn.config(state='normal')
            
        except Exception as e:
            self.log_status(f"\n✗ ERROR: {str(e)}", 'error')
            messagebox.showerror("Conversion Error", str(e))
    
    def save_output(self):
        if not self.converted_wb:
            messagebox.showerror("Error", "No conversion to save")
            return
        
        self.log_status("\nAttempting to save output...", 'info')
        
        saved = False
        while not saved:
            try:
                self.converted_wb.save(self.input_file)
                self.converted_wb.close()
                saved = True
                self.log_status(f"✓ Output saved to: {os.path.basename(self.input_file)}", 'success')
                
                messagebox.showinfo(
                    "Save Complete",
                    f"Successfully saved converted file:\n{os.path.basename(self.input_file)}\n\n"
                    + (f"⚠ {self.conversion_result['unmatched_count']} unmatched rows highlighted in yellow." 
                       if self.conversion_result['unmatched_count'] > 0 else "All rows converted successfully!")
                )
                
                self.save_btn.config(state='disabled')
                self.converted_wb = None
                self.conversion_result = None
                
            except PermissionError:
                response = messagebox.askretrycancel(
                    "File Locked",
                    f"The file '{os.path.basename(self.input_file)}' is currently open in Excel.\n\n"
                    "Please close it in Excel and click 'Retry'."
                )
                
                if not response:
                    self.converted_wb.close()
                    self.log_status("✗ Save cancelled - file not saved", 'error')
                    self.save_btn.config(state='disabled')
                    return
    
    def load_time_fees_table(self, time_fees_file):
        self.log_status(f"Loading time entries from {os.path.basename(time_fees_file)}...", 'info')
        
        wb = openpyxl.load_workbook(time_fees_file)
        
        if 'Table 1' not in wb.sheetnames:
            raise ValueError(f"Sheet 'Table 1' not found in {time_fees_file}")
        
        ws = wb['Table 1']
        
        header_row = None
        for row_num in range(1, min(10, ws.max_row + 1)):
            cell_value = ws.cell(row_num, 1).value
            if cell_value and str(cell_value).lower() == 'date':
                header_row = row_num
                break
        
        if header_row is None:
            raise ValueError("Could not find header row with 'Date' column")
        
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col).value
            if header:
                headers.append(str(header).strip())
            else:
                headers.append(None)
        
        num_headers = len([h for h in headers if h])
        
        if num_headers == 8:
            self.log_status("Detected 8-column layout - merging Staff columns...", 'info')
            
            staff_col = None
            surname_col = None
            
            for i, h in enumerate(headers):
                if h and 'Staff' in h:
                    staff_col = i + 1
                    surname_col = i + 2
                    break
            
            if staff_col and surname_col:
                for row_num in range(header_row + 1, ws.max_row + 1):
                    first_name = ws.cell(row_num, staff_col).value
                    surname = ws.cell(row_num, surname_col).value
                    
                    if first_name and surname:
                        full_name = f"{first_name} {surname}"
                        ws.cell(row_num, staff_col).value = full_name
                    elif surname:
                        ws.cell(row_num, staff_col).value = surname
                
                ws.delete_cols(surname_col, 1)
                self.log_status("Merged Staff columns successfully", 'success')
        
        data_rows = []
        for row_num in range(header_row + 1, ws.max_row + 1):
            # Load ALL rows including empty and duplicate headers
            row_data = {
                'row_num': row_num,
                'date': ws.cell(row_num, 1).value,
                'status': ws.cell(row_num, 2).value,
                'type': ws.cell(row_num, 3).value,
                'description': ws.cell(row_num, 4).value,
                'staff': ws.cell(row_num, 5).value,
                'hrs_qty': ws.cell(row_num, 6).value,
                'amount': ws.cell(row_num, 7).value
            }
            data_rows.append(row_data)
        
        self.log_status(f"Loaded {len(data_rows)} time entries", 'success')
        
        return wb, ws, data_rows
    
    def parse_amount(self, amount_str):
        if amount_str is None:
            return None
        
        if isinstance(amount_str, (int, float)):
            return float(amount_str)
        
        amount_str = str(amount_str).strip()
        amount_str = amount_str.replace('£', '').replace(',', '').strip()
        
        try:
            return float(amount_str)
        except ValueError:
            return None
    
    def detect_activity_type(self, current_rate):
        best_match = None
        best_diff = float('inf')
        
        for activity_key, rate_list in self.rates_data['rates'].items():
            for rate in rate_list:
                if rate > 0:
                    diff = abs(current_rate - rate)
                    
                    if diff < best_diff:
                        best_diff = diff
                        best_match = activity_key
        
        if best_diff < 1.00:
            return best_match, best_diff
        else:
            return None, None
    
    def perform_conversion(self, data_rows, target_index):
        self.log_status(f"Converting rates to target case type...", 'info')
        self.log_status(f"Processing {len(data_rows)} rows...", 'info')
        
        converted_rows = []
        unmatched_count = 0
        unmatched_rows = []
        
        for row in data_rows:
            # Check if row is completely empty (no meaningful data)
            is_empty_row = not any([
                row.get('date'),
                row.get('description'),
                row.get('amount'),
                row.get('hrs_qty'),
                row.get('staff')
            ])
            
            amount_val = self.parse_amount(row['amount'])
            hrs_qty_raw = row['hrs_qty']
            
            # Parse hrs_qty to a number
            try:
                if hrs_qty_raw is None or hrs_qty_raw == '':
                    hrs_qty = None
                elif isinstance(hrs_qty_raw, (int, float)):
                    hrs_qty = float(hrs_qty_raw)
                else:
                    hrs_qty = float(str(hrs_qty_raw).strip())
            except (ValueError, TypeError):
                hrs_qty = None
            
            # Skip empty rows or rows without amount/hrs_qty entirely
            if is_empty_row or amount_val is None or hrs_qty is None or hrs_qty == 0:
                # Empty rows or rows without data - mark as matched (neutral) so they don't appear in review
                converted_rows.append({
                    **row,
                    'new_amount': row['amount'],
                    'activity_type': None,
                    'matched': True  # Changed to True so empty rows aren't flagged for review
                })
                continue
            
            current_rate = amount_val / hrs_qty
            
            activity_type, diff = self.detect_activity_type(current_rate)
            
            if activity_type:
                target_rate = self.rates_data['rates'][activity_type][target_index]
                new_amount = hrs_qty * target_rate
                
                amount_str = f"£{new_amount:.3f}" if new_amount % 0.01 >= 0.001 else f"£{new_amount:.2f}"
                
                converted_rows.append({
                    **row,
                    'new_amount': amount_str,
                    'activity_type': activity_type,
                    'matched': True,
                    'old_rate': current_rate,
                    'new_rate': target_rate
                })
            else:
                unmatched_count += 1
                unmatched_rows.append(row['row_num'])
                
                converted_rows.append({
                    **row,
                    'new_amount': row['amount'],
                    'activity_type': None,
                    'matched': False,
                    'old_rate': current_rate
                })
        
        self.log_status(f"Converted: {len(data_rows) - unmatched_count} rows", 'success')
        if unmatched_count > 0:
            self.log_status(f"UNMATCHED: {unmatched_count} rows", 'warning')
        
        return {
            'converted_rows': converted_rows,
            'unmatched_count': unmatched_count,
            'unmatched_rows': unmatched_rows
        }
    
    def write_output(self, wb, ws, conversion_result, target_case_type):
        self.log_status("Writing output...", 'info')
        
        if 'orig_timeFees' in wb.sheetnames:
            del wb['orig_timeFees']
        
        ws.title = 'orig_timeFees'
        
        new_ws = wb.create_sheet('Table 1', 0)
        
        headers = ['Date', 'Status', 'Type', 'Description', 'Staff', 'Hrs/Qty', 'Amount']
        header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col_num, header in enumerate(headers, 1):
            cell = new_ws.cell(1, col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='top')
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        left_aligned = Alignment(horizontal='left', vertical='top')
        
        row_num = 2
        for conv_row in conversion_result['converted_rows']:
            date_cell = new_ws.cell(row_num, 1)
            date_cell.value = conv_row['date']
            date_cell.number_format = 'DD/MM/YYYY'
            date_cell.alignment = left_aligned
            
            status_cell = new_ws.cell(row_num, 2)
            status_cell.value = conv_row['status']
            status_cell.alignment = left_aligned
            
            type_cell = new_ws.cell(row_num, 3)
            type_cell.value = conv_row['type']
            type_cell.alignment = left_aligned
            
            desc_cell = new_ws.cell(row_num, 4)
            desc_cell.value = conv_row['description']
            desc_cell.alignment = left_aligned
            
            staff_cell = new_ws.cell(row_num, 5)
            staff_cell.value = conv_row['staff']
            staff_cell.alignment = left_aligned
            
            qty_cell = new_ws.cell(row_num, 6)
            qty_cell.value = conv_row['hrs_qty']
            qty_cell.alignment = left_aligned
            
            amount_cell = new_ws.cell(row_num, 7)
            amount_cell.value = conv_row['new_amount']
            amount_cell.alignment = left_aligned
            
            if not conv_row.get('matched', True):
                for col in range(1, 8):
                    new_ws.cell(row_num, col).fill = yellow_fill
            
            row_num += 1
        
        column_widths_chars = [14, 29, 14, 71, 29, 14, 14]
        for col_num, width in enumerate(column_widths_chars, 1):
            new_ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        self.log_status(f"Created new 'Table 1' sheet with {row_num - 2} rows", 'success')
        self.log_status(f"Renamed original to 'orig_timeFees'", 'info')
        if conversion_result['unmatched_count'] > 0:
            self.log_status(f"Highlighted {conversion_result['unmatched_count']} unmatched rows in YELLOW", 'warning')


def main():
    try:
        if HAS_DND:
            root = TkinterDnD.Tk()
        else:
            root = tk.Tk()
        app = RateConverterGUI(root)
        root.mainloop()
    except Exception as e:
        import traceback
        print(f"ERROR: {e}")
        print(traceback.format_exc())
        input("Press Enter to exit...")


if __name__ == "__main__":
    main()
