import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\thescoop\coding\rateConvert\Time & Fees.xlsx')
ws = wb['Table 1']

print("Checking rows 27, 28, 54, 79:\n")

for row_num in [27, 28, 54, 79]:
    date = ws.cell(row_num, 1).value
    status = ws.cell(row_num, 2).value
    type_val = ws.cell(row_num, 3).value
    desc = ws.cell(row_num, 4).value
    staff = ws.cell(row_num, 5).value
    hrs_qty = ws.cell(row_num, 6).value
    amount = ws.cell(row_num, 7).value
    
    print(f"Row {row_num}:")
    print(f"  date={date}, status={status}, type={type_val}")
    print(f"  desc={desc}, staff={staff}, hrs_qty={hrs_qty}, amount={amount}")
    
    # Check if empty
    is_empty = not any([date, desc, amount, hrs_qty, staff])
    print(f"  is_empty={is_empty}")
    print()
